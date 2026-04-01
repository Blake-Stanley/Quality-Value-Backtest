import openpyxl, numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import pandas as pd

OUT_DIR = Path(__file__).resolve().parent.parent / "Output"


def main():
    # ---- Load raw metrics ----
    df = pd.read_csv(OUT_DIR / "130_30_backtest_metrics.csv", index_col=0)

    portfolios = ["130/30 (EW)", "130/30 (VW)", "EW Long", "EW Short", "VW Long", "VW Short"]
    portfolios = [p for p in portfolios if p in df.index]

    rows = {p: df.loc[p].to_dict() for p in portfolios}

    def _f(d, k):
        v = d.get(k, np.nan)
        try: return float(v)
        except: return v

    metrics = [
        ("Period",               None,   lambda d: f"{str(d.get('start',''))[:10]} to {str(d.get('end',''))[:10]}"),
        ("Months",               None,   lambda d: int(d.get("months", 0))),
        ("Ann. Arith. Return",   "pct",  lambda d: _f(d,"ann_ret")),
        ("Ann. Geo. Return",     "pct",  lambda d: _f(d,"geo_ret")),
        ("Volatility",           "pct",  lambda d: _f(d,"ann_vol")),
        ("Sharpe Ratio",         "f2",   lambda d: _f(d,"sharpe")),
        ("t(mean)",              "f2",   lambda d: _f(d,"t_stat_ret")),
        ("p(mean)",              "f4",   lambda d: _f(d,"p_val_ret")),
        ("95% CI (Return)",      "ci",   lambda d: (_f(d,"ret_ci_lo"), _f(d,"ret_ci_hi"))),
        ("Max Drawdown",         "pct",  lambda d: _f(d,"max_dd")),
        ("CAPM Alpha",           "pct",  lambda d: _f(d,"capm_alpha")),
        ("Alpha Std. Error",     "pct",  lambda d: _f(d,"alpha_se")),
        ("t(alpha)",             "f2",   lambda d: _f(d,"alpha_t")),
        ("p(alpha)",             "f4",   lambda d: _f(d,"alpha_p")),
        ("95% CI (Alpha)",       "ci",   lambda d: (_f(d,"alpha_ci_lo"), _f(d,"alpha_ci_hi"))),
        ("CAPM Beta",            "f3",   lambda d: _f(d,"capm_beta")),
        ("t(beta)",              "f2",   lambda d: _f(d,"beta_t")),
        ("Sortino Ratio",        "f2",   lambda d: _f(d,"sortino")),
        ("Calmar Ratio",         "f2",   lambda d: _f(d,"calmar")),
        ("Worst Month",          "pct",  lambda d: _f(d,"worst_month")),
        ("% Positive Months",    "f1p",  lambda d: _f(d,"pct_positive")),
        ("Skewness",             "f2",   lambda d: _f(d,"skewness")),
        ("Excess Kurtosis",      "f2",   lambda d: _f(d,"kurtosis")),
    ]

    metric_dict = {m[0]: m for m in metrics}

    sections = {
        "Info":                    ["Period", "Months"],
        "Returns":                 ["Ann. Arith. Return", "Ann. Geo. Return", "Volatility",
                                    "Sharpe Ratio", "Max Drawdown"],
        "Risk":                    ["Sortino Ratio", "Calmar Ratio", "Worst Month",
                                    "% Positive Months", "Skewness", "Excess Kurtosis"],
        "Statistical Significance":["t(mean)", "p(mean)", "95% CI (Return)"],
        "CAPM":                    ["CAPM Alpha", "Alpha Std. Error", "t(alpha)", "p(alpha)",
                                    "95% CI (Alpha)", "CAPM Beta", "t(beta)"],
    }
    section_order = ["Info", "Returns", "Risk", "Statistical Significance", "CAPM"]

    def fmt(val, kind):
        if kind is None: return val
        if isinstance(val, float) and np.isnan(val): return "—"
        if kind == "pct":  return f"{val:.2%}"
        if kind == "f1p":  return f"{val:.1f}%"
        if kind == "f2":   return f"{val:.2f}"
        if kind == "f3":   return f"{val:.3f}"
        if kind == "f4":   return f"{val:.4f}"
        if kind == "ci":
            lo, hi = val
            if np.isnan(lo) or np.isnan(hi): return "—"
            return f"[{lo:.2%}, {hi:.2%}]"
        return val

    DARK_BLUE  = "1F3864"
    MED_BLUE   = "2E5FA3"
    LIGHT_BLUE = "D6E4F7"
    ALT_ROW    = "EEF4FB"
    WHITE      = "FFFFFF"

    col_colors = {
        "130/30 (EW)": "1F3864",
        "130/30 (VW)": "2E5FA3",
        "EW Long":     "2E75B6",
        "EW Short":    "C00000",
        "VW Long":     "375623",
        "VW Short":    "C55A11",
    }

    def thick(): return Side(style="medium")
    def hair():  return Side(style="hair")
    def thin():  return Side(style="thin")

    n_cols = len(portfolios) + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "130-30 Backtest Results"

    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = "130/30 Long-Short Equity ETF — Backtest Results"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = ("Composite z-score: Shareholder Yield + Gross Profitability + ROIC  |  "
               "130% long top-100 / 30% short bottom-100 (S&P 500 proxy)  |  "
               "Quarterly rebalancing, ±5 pp sector neutrality")
    c.font = Font(name="Calibri", italic=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=MED_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 5

    HR = 4
    ws.row_dimensions[HR].height = 22
    c = ws.cell(HR, 1, "Metric")
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(left=thick(), right=thick(), top=thick(), bottom=thick())

    for j, p in enumerate(portfolios, start=2):
        c = ws.cell(HR, j, p)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=col_colors.get(p, DARK_BLUE))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thick(), right=thick(), top=thick(), bottom=thick())

    current_row = HR + 1
    for sec in section_order:
        ws.merge_cells(f"A{current_row}:{get_column_letter(n_cols)}{current_row}")
        c = ws.cell(current_row, 1, sec.upper())
        c.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(top=thick(), bottom=thin(), left=thick(), right=thick())
        ws.row_dimensions[current_row].height = 15
        current_row += 1

        mnames = sections[sec]
        for mi, mname in enumerate(mnames):
            if mname not in metric_dict: continue
            label, kind, getter = metric_dict[mname]
            bg = WHITE if mi % 2 == 0 else ALT_ROW
            ws.row_dimensions[current_row].height = 15
            is_last = (mi == len(mnames) - 1)
            bot = thick() if is_last else hair()

            c = ws.cell(current_row, 1, label)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            c.border = Border(left=thick(), right=thin(), top=hair(), bottom=bot)

            for j, p in enumerate(portfolios, start=2):
                raw     = getter(rows[p])
                display = fmt(raw, kind)
                is_last_col = (j == len(portfolios) + 1)
                c = ws.cell(current_row, j, display)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(left=hair(), right=thick() if is_last_col else hair(),
                                  top=hair(), bottom=bot)
            current_row += 1

    ws.column_dimensions["A"].width = 26
    for j in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 20
    ws.freeze_panes = "B5"

    out_path = OUT_DIR / "130_30_Backtest_Table.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
