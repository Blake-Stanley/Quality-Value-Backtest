"""
factor_analysis.py — Factor exposure and return attribution analysis.

Reads from:
  Output/backtest_returns.csv   — monthly returns + S&P 500
  Data/ff5_plus_mom.dta         — Fama-French 5 factors + momentum

Outputs:
  Output/Factor_Analysis.xlsx   — styled Excel table
  Output/Charts/factor_correlations.png
  Output/Charts/factor_loadings.png
  Output/Charts/rolling_factor_corr.png
  Output/Charts/annual_returns_vs_sp500.png
  Output/Charts/cumulative_factors.png

Usage:
    python Code/factor_analysis.py
"""

import warnings
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from scipy import stats

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT      = Path(__file__).resolve().parent.parent
OUT_DIR   = ROOT / "Output"
DATA_DIR  = ROOT / "Data"

FF_FILE   = DATA_DIR / "ff5_plus_mom.dta"
RET_FILE  = OUT_DIR / "backtest_returns.csv"

# ── theme — imported from make_plots so colours stay in sync ──────────────────
from make_plots import (
    SLIDE_BG, AXES_BG, NAVY,
    C_LONG, C_SHORT, C_STRAT, C_SP500, C_PURPLE, C_LEVLONG, C_VW_STRAT,
    CHARTS_DIR, EW_DIR,
    _apply_theme, _pct_fmt, _save,
)

# extra colours used only in this file
C_RMW = "#2E8B57"
C_HML = "#7030A0"
C_MOM = "#C55A11"
C_CMA = "#4472C4"

# ── Excel theme (matches make_table.py) ───────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E5FA3"
LIGHT_BLUE = "D6E4F7"
ALT_ROW    = "EEF4FB"
WHITE      = "FFFFFF"

FACTOR_LABELS = {
    "mktrf": "Mkt-RF",
    "hml":   "HML (Value)",
    "rmw":   "RMW (Profitability)",
    "cma":   "CMA (Investment)",
    "umd":   "UMD (Momentum)",
    "smb":   "SMB (Size)",
}
FACTORS = ["mktrf", "hml", "rmw", "cma", "umd", "smb"]
FACTOR_COLORS = {
    "mktrf": C_STRAT,
    "hml":   C_HML,
    "rmw":   C_RMW,
    "cma":   C_CMA,
    "umd":   C_MOM,
    "smb":   "#888888",
}


# ── helpers ───────────────────────────────────────────────────────────────────
def _ols(y, X_cols, data):
    """OLS regression; returns dict of coef, se, t, p, r2, alpha_ann."""
    d = data.dropna(subset=[y] + X_cols)
    Y = d[y].values
    X = np.column_stack([np.ones(len(d))] + [d[c].values for c in X_cols])
    n, k = X.shape
    b, _, _, _ = np.linalg.lstsq(X, Y, rcond=None)
    resid = Y - X @ b
    s2 = (resid @ resid) / (n - k)
    cov = s2 * np.linalg.inv(X.T @ X)
    se = np.sqrt(np.diag(cov))
    t  = b / se
    p  = 2 * stats.t.sf(np.abs(t), df=n - k)
    ss_tot = ((Y - Y.mean()) ** 2).sum()
    ss_res = (resid ** 2).sum()
    r2 = 1 - ss_res / ss_tot
    names = ["alpha"] + X_cols
    result = {nm: {"coef": b[i], "se": se[i], "t": t[i], "p": p[i]}
              for i, nm in enumerate(names)}
    result["r2"] = r2
    result["n"]  = n
    result["alpha_ann"] = b[0] * 12
    result["alpha_ann_se"] = se[0] * np.sqrt(12)
    return result


# ── data loading ──────────────────────────────────────────────────────────────
def load_data():
    ret = pd.read_csv(RET_FILE)
    ret["date"] = pd.to_datetime(ret["date"])
    ret = ret.set_index("date")

    ff = pd.read_stata(FF_FILE)
    ff["date"] = pd.to_datetime(ff["dateff"]).dt.to_period("M").dt.to_timestamp()
    ff = ff.set_index("date")[["mktrf", "smb", "hml", "rmw", "cma", "umd", "rf"]]

    # merge on month-start aligned index
    ret.index = ret.index.to_period("M").to_timestamp()
    ff.index  = ff.index.to_period("M").to_timestamp()

    m = ret.join(ff, how="inner")
    # excess returns over rf
    m["ew_excess"] = m["ew_mkt_neutral"] - m["rf"]
    m["vw_excess"] = m["vw_mkt_neutral"] - m["rf"]
    return m


# ── analysis ──────────────────────────────────────────────────────────────────
def compute_factor_correlations(m):
    results = {}
    for series, label in [("ew_mkt_neutral", "EW Strategy"), ("vw_mkt_neutral", "VW Strategy")]:
        results[label] = {f: m[series].corr(m[f]) for f in FACTORS}
    return results


def compute_regressions(m):
    """Run FF5+Mom regression for EW and VW strategies."""
    regs = {}
    reg_factors = ["mktrf", "hml", "rmw", "cma", "umd"]
    for series, label in [("ew_excess", "EW Strategy"), ("vw_excess", "VW Strategy")]:
        regs[label] = _ols(series, reg_factors, m)
    return regs


def compute_annual(m):
    """Annual strategy and S&P 500 returns."""
    ann = {}
    for col in ["ew_mkt_neutral", "vw_mkt_neutral", "sp500"]:
        ann[col] = m[col].dropna().resample("YE").apply(lambda x: (1 + x).prod() - 1)
    return ann


def compute_rolling_corr(m, window=36):
    """Rolling correlations of EW strategy with key factors."""
    key_factors = ["mktrf", "rmw", "hml", "umd"]
    result = {}
    for f in key_factors:
        result[f] = m["ew_mkt_neutral"].rolling(window).corr(m[f])
    return result


# ══════════════════════════════════════════════════════════════════════════════
# CHARTS
# ══════════════════════════════════════════════════════════════════════════════

def plot_factor_correlations(corr_dict, ew_only=False):
    """Side-by-side bar chart of EW and VW correlations with each factor."""
    fig, ax = plt.subplots(figsize=(10, 5))

    labels = [FACTOR_LABELS[f] for f in FACTORS]
    x = np.arange(len(FACTORS))

    ew_vals = [corr_dict["EW Strategy"][f] for f in FACTORS]

    if ew_only:
        bars_ew = ax.bar(x, ew_vals, 0.55, label="EW Strategy",
                         color=C_STRAT, alpha=0.85, zorder=3)
        all_bars = list(bars_ew)
    else:
        vw_vals = [corr_dict["VW Strategy"][f] for f in FACTORS]
        bars_ew = ax.bar(x - 0.175, ew_vals, 0.35, label="EW Strategy",
                         color=C_STRAT, alpha=0.85, zorder=3)
        bars_vw = ax.bar(x + 0.175, vw_vals, 0.35, label="VW Strategy",
                         color=C_VW_STRAT, alpha=0.75, zorder=3)
        all_bars = list(bars_ew) + list(bars_vw)

    ax.axhline(0, color=NAVY, linewidth=0.8, alpha=0.6)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel("Pearson Correlation")
    ax.set_title("Strategy Correlation with Fama-French Factors", fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, axis="y", alpha=0.3, zorder=0)
    ax.set_ylim(-0.5, 0.6)

    for bar in all_bars:
        h = bar.get_height()
        va = "bottom" if h >= 0 else "top"
        offset = 0.01 if h >= 0 else -0.01
        ax.text(bar.get_x() + bar.get_width() / 2, h + offset,
                f"{h:.2f}", ha="center", va=va, fontsize=7.5, color=NAVY)

    _apply_theme(fig, [ax])
    fig.tight_layout()
    _save(fig, "factor_correlations.png", EW_DIR if ew_only else None)


def _draw_loadings_panel(ax, label, reg):
    reg_factors = ["mktrf", "hml", "rmw", "cma", "umd"]
    labels = [FACTOR_LABELS[f] for f in reg_factors]
    coefs = [reg[f]["coef"] for f in reg_factors]
    ses   = [reg[f]["se"]   for f in reg_factors]
    y     = np.arange(len(reg_factors))
    bar_colors = [C_LONG if c >= 0 else C_SHORT for c in coefs]
    ax.barh(y, coefs, xerr=ses, color=bar_colors, alpha=0.8,
            error_kw=dict(ecolor=NAVY, capsize=4, alpha=0.7), zorder=3)
    ax.axvline(0, color=NAVY, linewidth=0.8, alpha=0.6)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel("Factor Loading (β)")
    alpha_ann = reg["alpha_ann"]
    alpha_se  = reg["alpha_ann_se"]
    alpha_t   = reg["alpha"]["t"]
    r2        = reg["r2"]
    stars = ("***" if abs(alpha_t) > 2.576
             else "**" if abs(alpha_t) > 1.960
             else "*"  if abs(alpha_t) > 1.645
             else "")
    subtitle = (f"Ann. Alpha = {alpha_ann:+.1%}{stars}  (SE = {alpha_se:.1%})   "
                f"R² = {r2:.3f}")
    ax.set_title(f"{label} — FF5 Factor Loadings\n{subtitle}", fontsize=10)
    ax.grid(True, axis="x", alpha=0.3, zorder=0)


def plot_factor_loadings(regs, ew_only=False):
    """Horizontal bar chart of FF5 regression loadings for EW (and VW)."""
    if ew_only:
        fig, ax = plt.subplots(figsize=(8, 5))
        _draw_loadings_panel(ax, "EW Strategy", regs["EW Strategy"])
        all_axes = [ax]
    else:
        fig, axes = plt.subplots(1, 2, figsize=(13, 5))
        _draw_loadings_panel(axes[0], "EW Strategy", regs["EW Strategy"])
        _draw_loadings_panel(axes[1], "VW Strategy", regs["VW Strategy"])
        all_axes = list(axes)

    _apply_theme(fig, all_axes)
    fig.tight_layout(pad=2.0)
    _save(fig, "factor_loadings.png", EW_DIR if ew_only else None)


def plot_rolling_factor_corr(rolling_corr, window=36, ew_only=False):
    """Rolling correlation of EW strategy with key factors."""
    factor_map = {
        "mktrf": ("Mkt-RF", C_STRAT),
        "rmw":   ("RMW (Profitability)", C_RMW),
        "hml":   ("HML (Value)", C_HML),
        "umd":   ("UMD (Momentum)", C_MOM),
    }
    fig, ax = plt.subplots(figsize=(12, 5))

    for f, (lbl, color) in factor_map.items():
        s = rolling_corr[f].dropna()
        ax.plot(s.index, s.values, linewidth=1.6, label=lbl, color=color)

    ax.axhline(0, color=NAVY, linewidth=0.8, linestyle="--", alpha=0.5)
    ax.set_ylabel(f"{window}-Month Rolling Correlation")
    ax.set_title(f"Rolling {window}-Month Factor Correlations — EW Strategy", fontsize=11)
    ax.legend(fontsize=9, ncol=2)
    ax.grid(True, alpha=0.3)
    ax.set_ylim(-1, 1)

    _apply_theme(fig, [ax])
    fig.tight_layout()
    _save(fig, "rolling_factor_corr.png", EW_DIR if ew_only else None)


def plot_annual_returns(ann, ew_only=False):
    """Grouped bar chart: EW strategy vs (VW strategy) vs S&P 500 per year."""
    ew = ann["ew_mkt_neutral"]
    vw = ann["vw_mkt_neutral"]
    sp = ann["sp500"]

    idx = ew.index.union(sp.index)
    if not ew_only:
        idx = idx.union(vw.index)
    idx = idx[idx.year >= 1975]

    years = [i.year for i in idx]
    x = np.arange(len(years))

    ew_v = [ew.get(i, np.nan) for i in idx]
    sp_v = [sp.get(i, np.nan) for i in idx]

    fig, ax = plt.subplots(figsize=(16, 6))

    if ew_only:
        width = 0.38
        ax.bar(x - width / 2, ew_v, width, label="EW Strategy", color=C_STRAT, alpha=0.85, zorder=3)
        ax.bar(x + width / 2, sp_v, width, label="S&P 500",     color=C_SP500, alpha=0.75, zorder=3)
        title = "Annual Returns: EW Strategy vs S&P 500"
    else:
        vw_v  = [vw.get(i, np.nan) for i in idx]
        width = 0.28
        ax.bar(x - width, ew_v, width, label="EW Strategy", color=C_STRAT,    alpha=0.85, zorder=3)
        ax.bar(x,         vw_v, width, label="VW Strategy", color=C_VW_STRAT, alpha=0.75, zorder=3)
        ax.bar(x + width, sp_v, width, label="S&P 500",     color=C_SP500,    alpha=0.75, zorder=3)
        title = "Annual Returns: EW Strategy vs VW Strategy vs S&P 500"

    ax.axhline(0, color=NAVY, linewidth=0.8, alpha=0.6)
    ax.set_xticks(x)
    ax.set_xticklabels(years, rotation=90, fontsize=7)
    _pct_fmt(ax)
    ax.set_ylabel("Annual Return")
    ax.set_title(title, fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, axis="y", alpha=0.3, zorder=0)

    _apply_theme(fig, [ax])
    fig.tight_layout()
    _save(fig, "annual_returns_vs_sp500.png", EW_DIR if ew_only else None)


def plot_cumulative_factors(m, ew_only=False):
    """Cumulative growth of strategy and key FF factors on same chart."""
    fig, ax = plt.subplots(figsize=(12, 6))

    series = [
        ("ew_mkt_neutral", "EW Strategy",          C_STRAT,   2.0, "-"),
        ("mktrf",          "Mkt-RF",                C_SP500,   1.4, "--"),
        ("rmw",            "RMW (Profitability)",   C_RMW,     1.4, "--"),
        ("hml",            "HML (Value)",           C_HML,     1.4, ":"),
        ("umd",            "UMD (Momentum)",        C_MOM,     1.2, "-."),
    ]

    for col, lbl, color, lw, ls in series:
        s = m[col].dropna()
        cum = (1 + s).cumprod()
        ax.plot(cum.index, cum.values, linewidth=lw, label=lbl,
                color=color, linestyle=ls)

    ax.set_yscale("log")
    ax.axhline(1, color=NAVY, linewidth=0.6, linestyle=":", alpha=0.4)
    ax.set_ylabel("Cumulative Return (log scale, $1 invested)")
    ax.set_title("Cumulative Performance: Strategy vs Key Fama-French Factors", fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3, which="both")

    _apply_theme(fig, [ax])
    fig.tight_layout()
    _save(fig, "cumulative_factors.png", EW_DIR if ew_only else None)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL TABLE
# ══════════════════════════════════════════════════════════════════════════════

def _thick(): return Side(style="medium")
def _hair():  return Side(style="hair")
def _thin():  return Side(style="thin")


def _cell(ws, row, col, value, bold=False, size=10, color="000000", bg=WHITE,
          halign="center", indent=0, border=None):
    c = ws.cell(row, col, value)
    c.font = Font(name="Calibri", bold=bold, size=size, color=color)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical="center",
                             wrap_text=True, indent=indent)
    if border:
        c.border = border
    return c


def build_excel(corr_dict, regs, m):
    reg_factors = ["mktrf", "hml", "rmw", "cma", "umd"]
    series_labels = ["EW Strategy", "VW Strategy"]
    col_colors = {"EW Strategy": DARK_BLUE, "VW Strategy": MED_BLUE}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factor Analysis"

    n_cols = len(series_labels) + 1

    # ── title rows ────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = "Market Neutral Long-Short Equity ETF — Factor Analysis"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    start = m.index.min().strftime("%b %Y")
    end   = m.index.max().strftime("%b %Y")
    c.value = f"Fama-French 5 Factor + Momentum attribution  |  {start} – {end}  |  Monthly returns"
    c.font = Font(name="Calibri", italic=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=MED_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 5

    # ── header row ────────────────────────────────────────────────────────────
    HR = 4
    ws.row_dimensions[HR].height = 22
    _cell(ws, HR, 1, "Metric", bold=True, size=11, color="FFFFFF", bg=DARK_BLUE,
          halign="left", indent=1,
          border=Border(left=_thick(), right=_thick(), top=_thick(), bottom=_thick()))
    for j, lbl in enumerate(series_labels, start=2):
        _cell(ws, HR, j, lbl, bold=True, size=10, color="FFFFFF",
              bg=col_colors[lbl], halign="center",
              border=Border(left=_thick(), right=_thick(), top=_thick(), bottom=_thick()))

    row = HR + 1

    def _section_header(label):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row, 1, label.upper())
        c.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = Border(top=_thick(), bottom=_thin(), left=_thick(), right=_thick())
        ws.row_dimensions[row].height = 15
        row += 1

    def _data_row(label, values, fmt_fn, is_last=False):
        nonlocal row
        bg = WHITE if (row % 2 == 0) else ALT_ROW
        bot = _thick() if is_last else _hair()
        ws.row_dimensions[row].height = 15
        c = ws.cell(row, 1, label)
        c.font = Font(name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border = Border(left=_thick(), right=_thin(), top=_hair(), bottom=bot)
        for j, v in enumerate(values, start=2):
            is_last_col = (j == n_cols)
            c = ws.cell(row, j, fmt_fn(v))
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(left=_hair(),
                              right=_thick() if is_last_col else _hair(),
                              top=_hair(), bottom=bot)
        row += 1

    # ── SECTION 1: Factor Correlations ────────────────────────────────────────
    _section_header("Factor Correlations (Pearson, full sample)")
    for i, f in enumerate(FACTORS):
        vals = [corr_dict[lbl][f] for lbl in series_labels]
        is_last = (i == len(FACTORS) - 1)
        _data_row(FACTOR_LABELS[f], vals, lambda v: f"{v:.3f}", is_last=is_last)

    # ── SECTION 2: FF5+Mom Regression ─────────────────────────────────────────
    _section_header("FF5 + Momentum Regression (excess returns on excess market)")

    # annualised alpha
    vals = [regs[lbl]["alpha_ann"] for lbl in series_labels]
    _data_row("Ann. Alpha", vals, lambda v: f"{v:.2%}")

    vals = [regs[lbl]["alpha_ann_se"] for lbl in series_labels]
    _data_row("Alpha Std. Error (ann.)", vals, lambda v: f"{v:.2%}")

    vals = [regs[lbl]["alpha"]["t"] for lbl in series_labels]
    _data_row("t(alpha)", vals, lambda v: f"{v:.2f}")

    vals = [regs[lbl]["alpha"]["p"] for lbl in series_labels]
    _data_row("p(alpha)", vals, lambda v: f"{v:.4f}")

    vals = [regs[lbl]["r2"] for lbl in series_labels]
    _data_row("R²", vals, lambda v: f"{v:.3f}")

    vals = [regs[lbl]["n"] for lbl in series_labels]
    _data_row("Observations", vals, lambda v: str(int(v)))

    # factor loadings
    for i, f in enumerate(reg_factors):
        lbl_f = FACTOR_LABELS[f]
        vals_b = [regs[lbl][f]["coef"] for lbl in series_labels]
        _data_row(f"{lbl_f}  β", vals_b, lambda v: f"{v:.3f}")

        vals_t = [regs[lbl][f]["t"] for lbl in series_labels]
        is_last = (i == len(reg_factors) - 1)
        _data_row(f"{lbl_f}  t", vals_t, lambda v: f"{v:.2f}", is_last=is_last)

    # ── SECTION 3: Annual Return Summary ──────────────────────────────────────
    _section_header("Annual Returns")
    ann_ew = m["ew_mkt_neutral"].resample("YE").apply(lambda x: (1 + x).prod() - 1)
    ann_vw = m["vw_mkt_neutral"].resample("YE").apply(lambda x: (1 + x).prod() - 1)
    ann_sp = m["sp500"].resample("YE").apply(lambda x: (1 + x).prod() - 1)

    years = sorted(set(ann_ew.index) | set(ann_vw.index))
    years = [y for y in years if y.year >= 1975]

    # add a "Year" column by widening to 3 data cols
    # rebuild header to include S&P 500
    # use a simple approach: encode year in metric label
    for i, yr in enumerate(years):
        y_str = str(yr.year)
        ew_v = ann_ew.get(yr, np.nan)
        vw_v = ann_vw.get(yr, np.nan)
        sp_v = ann_sp.get(yr, np.nan)
        sp_str = f"{sp_v:.2%}" if not np.isnan(sp_v) else "—"
        label = f"{y_str}   [S&P 500: {sp_str}]"
        vals = [ew_v, vw_v]
        is_last = (i == len(years) - 1)
        _data_row(label, vals, lambda v: "—" if np.isnan(v) else f"{v:.2%}", is_last=is_last)

    # ── column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 38
    for j in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.freeze_panes = "B5"

    out_path = OUT_DIR / "Factor_Analysis.xlsx"
    wb.save(out_path)
    print(f"  Saved: {out_path.name}")


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading data...")
    m = load_data()
    print(f"  {len(m)} months ({m.index.min().strftime('%b %Y')} – {m.index.max().strftime('%b %Y')})")

    print("Computing factor correlations...")
    corr_dict = compute_factor_correlations(m)

    print("Running FF5+Mom regressions...")
    regs = compute_regressions(m)

    print("Computing annual returns...")
    ann = compute_annual(m)

    print("Computing rolling correlations...")
    rolling_corr = compute_rolling_corr(m, window=36)

    print("Generating charts...")
    for ew_only in (False, True):
        plot_factor_correlations(corr_dict, ew_only=ew_only)
        plot_factor_loadings(regs, ew_only=ew_only)
        plot_rolling_factor_corr(rolling_corr, window=36, ew_only=ew_only)
        plot_annual_returns(ann, ew_only=ew_only)
        plot_cumulative_factors(m, ew_only=ew_only)

    print("Building Excel table...")
    build_excel(corr_dict, regs, m)

    print("Done.")


if __name__ == "__main__":
    main()
