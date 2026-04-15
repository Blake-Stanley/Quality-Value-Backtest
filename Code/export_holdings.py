"""
Generate the latest long/short holdings snapshot for the market neutral strategy.

Exports a styled Excel workbook (three sheets: All Holdings, Long Book, Short Book)
using the same visual style as make_table.py.

Long book columns  : Shareholder Yield, Gross Profitability, ROIC (quality/value factors)
Short book columns : FCF Yield, Accruals, EV/EBIT, Net Ext. Fin., F-Score, Leverage
                     (failure-model factors inspired by Empirical Research Partners)
"""

from pathlib import Path
import sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import backtest as bt

OUTPUT_FILE = bt.OUT_DIR / "holdings_snapshot.xlsx"

# ── Color palette (matches make_table.py) ─────────────────────────────────────
DARK_BLUE   = "1F3864"
MED_BLUE    = "2E5FA3"
LIGHT_BLUE  = "D6E4F7"
ALT_ROW     = "EEF4FB"
WHITE       = "FFFFFF"
DARK_GREEN  = "375623"
MED_GREEN   = "548235"
LIGHT_GREEN = "E2EFDA"
ALT_GREEN   = "F0F7EE"
DARK_RED    = "C00000"
MED_RED     = "D9534F"
LIGHT_RED   = "FCE4D6"
ALT_RED     = "FEF2EE"

# ── Long-signal factor column names ───────────────────────────────────────────
_LONG_FACTOR_Z_COLS = ["Shareholder Yield Z", "Gross Profitability Z", "ROIC Z"]
_LONG_RAW_COLS      = ["Shareholder Yield (TTM)", "Gross Profitability", "ROIC", "P/E (TTM)"]

# ── Short-signal factor column names ──────────────────────────────────────────
_SHORT_FACTOR_Z_COLS = ["FCF Yield Z", "Accruals Z", "EV/EBIT Z", "NEF Z", "F-Score Z", "Leverage Z"]
_SHORT_RAW_COLS      = ["FCF Yield (TTM)", "Accruals", "Net Ext. Fin.", "F-Score", "Leverage"]

# ── Column group definitions ───────────────────────────────────────────────────
# All Holdings: shows both sets of factor z-scores (cross-book entries show "—")
_ALL_GROUPS = [
    ("POSITION",               DARK_BLUE,  ["Book", "Rank", "Book Weight"]),
    ("IDENTITY",               MED_BLUE,   ["Ticker", "Company", "Sector", "PERMNO"]),
    ("DATES",                  "4472C4",   ["Rebalance Month", "Signal Month", "Financials Through"]),
    ("COMPOSITE SCORE",        DARK_GREEN, ["Composite Z"]),
    ("LONG FACTORS (Z-SCORE)", MED_GREEN,  _LONG_FACTOR_Z_COLS),
    ("SHORT FACTORS (Z-SCORE)", DARK_RED,  _SHORT_FACTOR_Z_COLS),
]

# Long Book: quality/value factors with raw values
_LONG_GROUPS = [
    ("POSITION",        DARK_BLUE,  ["Rank", "Book Weight"]),
    ("IDENTITY",        MED_BLUE,   ["Ticker", "Company", "Sector", "PERMNO"]),
    ("DATES",           "4472C4",   ["Rebalance Month", "Signal Month", "Financials Through"]),
    ("COMPOSITE SCORE", DARK_GREEN, ["Composite Z"]),
    ("FACTOR Z-SCORES", MED_GREEN,  _LONG_FACTOR_Z_COLS),
    ("RAW FACTORS",     "70AD47",   _LONG_RAW_COLS),
]

# Short Book: failure-model factors with raw values
_SHORT_GROUPS = [
    ("POSITION",               DARK_BLUE,  ["Rank", "Book Weight"]),
    ("IDENTITY",               MED_BLUE,   ["Ticker", "Company", "Sector", "PERMNO"]),
    ("DATES",                  "4472C4",   ["Rebalance Month", "Signal Month", "Financials Through"]),
    ("COMPOSITE SCORE",        DARK_GREEN, ["Composite Z"]),
    ("SHORT FACTOR Z-SCORES",  DARK_RED,   _SHORT_FACTOR_Z_COLS),
    ("SHORT RAW FACTORS",      MED_RED,    _SHORT_RAW_COLS),
]

_COL_WIDTHS = {
    "Book": 8, "Rank": 6, "Book Weight": 11,
    "Ticker": 8, "Company": 28, "Sector": 16, "PERMNO": 8,
    "Rebalance Month": 14, "Signal Month": 13, "Financials Through": 15,
    "Composite Z": 11,
    # Long factor columns
    "Shareholder Yield Z": 15, "Gross Profitability Z": 17, "ROIC Z": 9,
    "Shareholder Yield (TTM)": 17, "Gross Profitability": 17, "ROIC": 9, "P/E (TTM)": 10,
    # Short factor z-score columns
    "FCF Yield Z": 10, "Accruals Z": 10, "EV/EBIT Z": 10,
    "NEF Z": 8, "F-Score Z": 10, "Leverage Z": 10,
    # Short raw columns
    "FCF Yield (TTM)": 14, "Accruals": 10, "Net Ext. Fin.": 13, "F-Score": 9, "Leverage": 10,
}

# Columns that should be formatted as percentages
_PCT_COLS = {
    "Shareholder Yield (TTM)", "Gross Profitability", "ROIC",
    "FCF Yield (TTM)", "Accruals", "Net Ext. Fin.", "Leverage",
    "Book Weight",
}
# Columns that should be formatted as rounded floats (z-scores etc.)
_FLOAT2_COLS = {
    "Composite Z",
    "Shareholder Yield Z", "Gross Profitability Z", "ROIC Z",
    "FCF Yield Z", "Accruals Z", "EV/EBIT Z", "NEF Z", "F-Score Z", "Leverage Z",
}


def _thick(): return Side(style="medium")
def _hair():  return Side(style="hair")
def _thin():  return Side(style="thin")


def _fmt_val(col: str, val):
    """Format a cell value for display based on its column."""
    try:
        is_nan = pd.isna(val)
    except (TypeError, ValueError):
        is_nan = False
    if is_nan:
        return "—"
    if col in _PCT_COLS:
        return f"{float(val):.2%}"
    if col in _FLOAT2_COLS:
        return round(float(val), 2)
    if col == "Rank":
        return int(val)
    if col == "F-Score":
        return round(float(val), 1)
    if col == "P/E (TTM)":
        fv = float(val)
        if np.isinf(fv) or np.isnan(fv):
            return "—"
        return round(fv, 1)
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m")
    return val


def _write_styled_sheet(ws, df: pd.DataFrame, title: str, subtitle: str,
                        column_groups: list, book_type: str = None):
    """Write a fully styled holdings table into an openpyxl worksheet."""
    all_cols = [c for _, _, cols in column_groups for c in cols if c in df.columns]
    df = df[all_cols].reset_index(drop=True)
    n_cols = len(all_cols)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Subtitle ───────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    c = ws["A2"]
    c.value = subtitle
    c.font = Font(name="Calibri", italic=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=MED_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # ── Row 3: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 5

    # ── Row 4: Column group headers ───────────────────────────────────────────
    ws.row_dimensions[4].height = 16
    col_idx = 1
    for grp_label, grp_color, grp_cols in column_groups:
        present = [c for c in grp_cols if c in df.columns]
        if not present:
            continue
        n = len(present)
        start_ltr = get_column_letter(col_idx)
        end_ltr   = get_column_letter(col_idx + n - 1)
        if n > 1:
            ws.merge_cells(f"{start_ltr}4:{end_ltr}4")
        cell = ws.cell(4, col_idx, grp_label)
        cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=grp_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(left=_thick(), right=_thick(), top=_thick(), bottom=_thin())
        col_idx += n

    # ── Row 5: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[5].height = 32
    col_idx = 1
    for _, grp_color, grp_cols in column_groups:
        present = [c for c in grp_cols if c in df.columns]
        for col_name in present:
            cell = ws.cell(5, col_idx, col_name)
            cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor=grp_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thick())
            col_idx += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    n_data = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = i + 6
        book = row.get("Book", book_type or "")
        if book == "Long":
            bg = LIGHT_GREEN if i % 2 == 0 else ALT_GREEN
        elif book == "Short":
            bg = LIGHT_RED if i % 2 == 0 else ALT_RED
        else:
            bg = WHITE if i % 2 == 0 else ALT_ROW

        ws.row_dimensions[excel_row].height = 14
        is_last_row = (i == n_data - 1)
        bot = _thick() if is_last_row else _hair()

        for j, col_name in enumerate(all_cols, start=1):
            val = _fmt_val(col_name, row[col_name])
            cell = ws.cell(excel_row, j, val)
            cell.font  = Font(name="Calibri", size=9)
            cell.fill  = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left" if col_name in ("Company", "Sector") else "center",
                vertical="center",
            )
            cell.border = Border(
                left=_thick() if j == 1 else _hair(),
                right=_thick() if j == n_cols else _hair(),
                top=_hair(),
                bottom=bot,
            )

    # ── Column widths & freeze ────────────────────────────────────────────────
    col_idx = 1
    for _, _, grp_cols in column_groups:
        for col_name in grp_cols:
            if col_name in df.columns:
                ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 12)
                col_idx += 1

    ws.freeze_panes = "A6"


def _latest_component_snapshot(signal_df: pd.DataFrame) -> pd.DataFrame:
    df = signal_df.copy()
    df["signal_month"] = df["signal_avail"].dt.to_period("M").dt.to_timestamp()
    df.sort_values(["PERMNO", "signal_month", "datadate"], inplace=True)
    keep_cols = [
        "PERMNO", "signal_month", "datadate", "tic", "conm",
        # Long signal raw + z
        "sh_yield", "gross_prof", "roic",
        "sh_yield_z", "gross_prof_z", "roic_z",
        "pe_ratio",
        # Short signal raw (ev_ebit not exported as raw, only z)
        "fcf_yield", "accruals", "nef", "f_score", "leverage",
        # Short signal z
        "fcf_yield_z", "accruals_z", "ev_ebit_z", "nef_z", "f_score_z", "leverage_z",
    ]
    # Only keep columns that actually exist in the cache
    keep_cols = [c for c in keep_cols if c in df.columns]
    return df[keep_cols].drop_duplicates(subset=["PERMNO", "signal_month"], keep="last")


def main():
    _cache_merged = bt.CACHE_DIR / "merged.parquet"
    _cache_signal = bt.CACHE_DIR / "signal_components.parquet"

    if _cache_merged.exists() and _cache_signal.exists():
        merged     = pd.read_parquet(_cache_merged,  engine="pyarrow")
        signal_raw = pd.read_parquet(_cache_signal,  engine="pyarrow")
        for col in ["month", "signal_source_month"]:
            if col in merged.columns:
                merged[col] = pd.to_datetime(merged[col])
        for col in ["signal_avail", "datadate"]:
            if col in signal_raw.columns:
                signal_raw[col] = pd.to_datetime(signal_raw[col])
    else:
        comp, crsp, _ = bt.load_data()
        signal_raw    = bt.build_signal(comp, include_components=True)
        signal_monthly = bt.resample_signal(signal_raw)
        crsp_clean    = bt.clean_crsp(crsp)
        merged        = bt.merge_and_form_portfolios(crsp_clean, signal_monthly)
        bt.CACHE_DIR.mkdir(exist_ok=True)
        merged.to_parquet(_cache_merged,  engine="pyarrow", index=False)
        signal_raw.to_parquet(_cache_signal, engine="pyarrow", index=False)

    if merged.empty:
        raise RuntimeError("No merged rows were produced; cannot build holdings.")

    latest_month = merged["month"].max()
    snapshot = merged[
        (merged["month"] == latest_month) & (merged["port"].isin(["long", "short"]))
    ].copy()
    if snapshot.empty:
        raise RuntimeError("No holdings found for the latest month.")

    snapshot["Book"] = snapshot["port"].str.title()
    snapshot["Book Weight"] = np.where(
        snapshot["Book"] == "Long",
        bt.LONG_WEIGHT / bt.N_LONG,
        (bt.LONG_WEIGHT - 1.0) / bt.N_SHORT,
    )

    # Composite Z: use the correct signal per book
    snapshot["Composite Z"] = np.where(
        snapshot["Book"] == "Long",
        snapshot[bt.SIGNAL_COL],
        snapshot[bt.SHORT_SIGNAL_COL],
    )

    # Rank each book by its own signal
    snapshot["Rank"] = np.nan
    long_mask  = snapshot["Book"] == "Long"
    short_mask = snapshot["Book"] == "Short"
    snapshot.loc[long_mask,  "Rank"] = snapshot.loc[long_mask,  bt.SIGNAL_COL].rank(ascending=False, method="first")
    snapshot.loc[short_mask, "Rank"] = snapshot.loc[short_mask, bt.SHORT_SIGNAL_COL].rank(ascending=False, method="first")
    snapshot["Rank"] = snapshot["Rank"].astype(int)

    # Merge in factor components from the signal cache
    components = _latest_component_snapshot(signal_raw)
    snapshot = snapshot.merge(
        components,
        left_on=["PERMNO", "signal_source_month"],
        right_on=["PERMNO", "signal_month"],
        how="left",
    )

    snapshot["Sector"]             = snapshot["sich"].map(bt._SIC_MAP).fillna("Industrials")
    snapshot["Rebalance Month"]    = snapshot["month"].dt.to_period("M").dt.to_timestamp()
    snapshot["Signal Month"]       = snapshot["signal_source_month"].dt.to_period("M").dt.to_timestamp()
    snapshot["Financials Through"] = snapshot["datadate"].dt.to_period("M").dt.to_timestamp()

    # Rename all factor columns to display names
    snapshot.rename(columns={
        "tic":          "Ticker",
        "conm":         "Company",
        # Long signal raw + z
        "sh_yield":     "Shareholder Yield (TTM)",
        "gross_prof":   "Gross Profitability",
        "roic":         "ROIC",
        "sh_yield_z":   "Shareholder Yield Z",
        "gross_prof_z": "Gross Profitability Z",
        "roic_z":       "ROIC Z",
        "pe_ratio":     "P/E (TTM)",
        # Short signal raw
        "fcf_yield":    "FCF Yield (TTM)",
        "accruals":     "Accruals",
        "nef":          "Net Ext. Fin.",
        "f_score":      "F-Score",
        "leverage":     "Leverage",
        # Short signal z
        "fcf_yield_z":  "FCF Yield Z",
        "accruals_z":   "Accruals Z",
        "ev_ebit_z":    "EV/EBIT Z",
        "nef_z":        "NEF Z",
        "f_score_z":    "F-Score Z",
        "leverage_z":   "Leverage Z",
    }, inplace=True)

    # For All Holdings sheet: blank out cross-book factor columns so each row
    # only shows factors relevant to its own signal
    for col in _LONG_FACTOR_Z_COLS + _LONG_RAW_COLS:
        if col in snapshot.columns:
            snapshot.loc[snapshot["Book"] == "Short", col] = np.nan
    for col in _SHORT_FACTOR_Z_COLS + _SHORT_RAW_COLS:
        if col in snapshot.columns:
            snapshot.loc[snapshot["Book"] == "Long", col] = np.nan

    snapshot = snapshot.sort_values(["Book", "Rank"]).reset_index(drop=True)

    # Split books (before blanking — each has its own full data since we already masked above)
    long_df  = snapshot[snapshot["Book"] == "Long"].copy()
    short_df = snapshot[snapshot["Book"] == "Short"].copy()

    rebal_str   = latest_month.strftime("%B %Y")
    title_all   = f"Market Neutral Long-Short Equity — Holdings Snapshot  ({rebal_str})"
    title_long  = f"Market Neutral Long-Short Equity — Long Book  ({rebal_str})"
    title_short = f"Market Neutral Long-Short Equity — Short Book  ({rebal_str})"
    subtitle = (
        "Long signal: Shareholder Yield + Gross Profitability + ROIC  |  "
        "Short signal: FCF Yield (neg, 1.5×) + Accruals + EV/EBIT + NEF + F-Score (neg) + Leverage (0.5×) + Gross Prof (neg, 0.5×)  |  "
        f"{bt.LONG_WEIGHT:.0%} long top-{bt.N_LONG} / w_short solved for {bt.TARGET_BETA:.2f} net beta "
        f"(Vasicek-adj, {bt.BETA_WINDOW}m trailing)  |  "
        f"Monthly rebalancing, ±{bt.SECTOR_TOL:.0%} sector neutrality"
    )

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "All Holdings"
    _write_styled_sheet(ws_all, snapshot, title_all, subtitle, _ALL_GROUPS)

    ws_long = wb.create_sheet("Long Book")
    _write_styled_sheet(ws_long, long_df, title_long, subtitle, _LONG_GROUPS, book_type="Long")

    ws_short = wb.create_sheet("Short Book")
    _write_styled_sheet(ws_short, short_df, title_short, subtitle, _SHORT_GROUPS, book_type="Short")

    wb.save(OUTPUT_FILE)
    print(f"Saved holdings snapshot to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
