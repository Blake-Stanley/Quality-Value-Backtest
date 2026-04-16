"""
Factor regression: regress EW market-neutral strategy returns on style factors.
Compounds daily factor returns to monthly, aligns with strategy monthly returns,
runs OLS, and outputs a styled Excel table.
"""

import numpy as np
import pandas as pd
import statsmodels.api as sm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "Data"
OUT_DIR  = BASE_DIR / "Output"

FACTORS = [
    "SmallSize",
    "CreditRisk",
    "Momentum",
    "Value",
    "Quality",
    "LowVolatility",
]

FACTOR_LABELS = {
    "SmallSize":     "Small Size",
    "CreditRisk":    "Credit Risk",
    "Momentum":      "Momentum",
    "Value":         "Value",
    "Quality":       "Quality",
    "LowVolatility": "Low Volatility",
}


# ── colours matching make_table.py ──────────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E5FA3"
LIGHT_BLUE = "D6E4F7"
ALT_ROW    = "EEF4FB"
WHITE      = "FFFFFF"
GREEN_SIG  = "E2EFDA"   # light green for significant rows
RED_HEX    = "C00000"


def thick(): return Side(style="medium")
def thin():  return Side(style="thin")
def hair():  return Side(style="hair")


def stars(p):
    if p < 0.01:  return "***"
    if p < 0.05:  return "**"
    return ""


def load_monthly_factors():
    """Load daily factor returns, compound to monthly."""
    df = pd.read_excel(DATA_DIR / "factor_returns.xlsx")
    df["Date"] = pd.to_datetime(df["Date"])
    df = df[["Date"] + FACTORS].copy()
    df = df.sort_values("Date")

    # year-month period for grouping
    df["ym"] = df["Date"].dt.to_period("M")

    # compound daily → monthly: prod(1+r) - 1
    # min_count=1 ensures months with all-NaN values stay NaN (not 1.0)
    monthly = (
        df.groupby("ym")[FACTORS]
        .apply(lambda g: (1 + g).prod(min_count=1) - 1)
    )
    monthly.index = monthly.index.to_timestamp()   # period → timestamp (month start)
    return monthly


def load_strategy_returns():
    """Load EW market-neutral monthly returns."""
    df = pd.read_csv(OUT_DIR / "backtest_returns.csv", parse_dates=["date"])
    df = df[["date", "ew_mkt_neutral"]].dropna().copy()
    df["ym"] = df["date"].dt.to_period("M").dt.to_timestamp()
    df = df.set_index("ym")[["ew_mkt_neutral"]]
    return df


def run_regression(y, X_raw):
    """OLS with statsmodels; returns result object."""
    X = sm.add_constant(X_raw)
    model = sm.OLS(y, X).fit(cov_type="HC3")   # heteroscedasticity-robust SEs
    return model


def build_excel(result, y, X_raw, out_path):
    FACTOR_COLS = X_raw.columns.tolist()
    n_obs    = int(result.nobs)
    r2       = result.rsquared
    adj_r2   = result.rsquared_adj
    f_stat   = result.fvalue
    f_pval   = result.f_pvalue

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Factor Regression"

    # ── title row ────────────────────────────────────────────────────────────
    n_data_cols = 5          # Factor | Coeff | Std Err | t-stat | p-value
    total_cols  = n_data_cols + 1   # +1 for sig column

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    c = ws["A1"]
    c.value = "EW Market-Neutral Strategy — Style Factor Regression"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── subtitle ─────────────────────────────────────────────────────────────
    start_date = X_raw.index.min().strftime("%b %Y")
    end_date   = X_raw.index.max().strftime("%b %Y")
    ws.merge_cells(f"A2:{get_column_letter(total_cols)}2")
    c = ws["A2"]
    c.value = (
        f"OLS regression of monthly EW strategy returns on style factor returns  |  "
        f"Sample: {start_date} – {end_date}  |  N = {n_obs}  |  "
        f"Robust standard errors (HC3)  |  ** p<0.05  *** p<0.01"
    )
    c.font = Font(name="Calibri", italic=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=MED_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # spacer
    ws.row_dimensions[3].height = 5

    # ── column header row ────────────────────────────────────────────────────
    HR = 4
    ws.row_dimensions[HR].height = 22
    headers = ["Factor", "Coefficient", "Std. Error", "t-Statistic", "p-Value", "Sig."]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(HR, j, h)
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center" if j > 1 else "left",
                                vertical="center")
        c.border = Border(left=thick(), right=thick(), top=thick(), bottom=thick())
        if j == 1:
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── intercept + factor rows ───────────────────────────────────────────────
    row_labels = ["Intercept"] + [FACTOR_LABELS[f] for f in FACTOR_COLS]
    param_names = ["const"] + FACTOR_COLS

    current_row = HR + 1

    # section header
    ws.merge_cells(f"A{current_row}:{get_column_letter(total_cols)}{current_row}")
    c = ws.cell(current_row, 1, "FACTOR LOADINGS")
    c.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(top=thick(), bottom=thin(), left=thick(), right=thick())
    ws.row_dimensions[current_row].height = 15
    current_row += 1

    for i, (label, pname) in enumerate(zip(row_labels, param_names)):
        coef   = result.params[pname]
        se     = result.bse[pname]
        tstat  = result.tvalues[pname]
        pval   = result.pvalues[pname]
        sig    = stars(pval)

        is_last = (i == len(row_labels) - 1)
        bg = WHITE if i % 2 == 0 else ALT_ROW
        # highlight significant factors (non-intercept)
        if pname != "const" and pval < 0.05:
            bg = GREEN_SIG if i % 2 == 0 else "D5E8D4"

        bot = thick() if is_last else hair()
        ws.row_dimensions[current_row].height = 15

        row_data = [
            (label,           "left",   None),
            (f"{coef:+.4f}",  "center", coef),
            (f"{se:.4f}",     "center", None),
            (f"{tstat:+.2f}", "center", tstat),
            (f"{pval:.4f}",   "center", None),
            (sig,             "center", None),
        ]

        for j, (val, align, _num) in enumerate(row_data, start=1):
            is_last_col = (j == total_cols)
            c = ws.cell(current_row, j, val)
            c.font = Font(name="Calibri", size=10,
                          bold=(pname != "const" and pval < 0.05))
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center",
                                    indent=2 if j == 1 else 0)
            c.border = Border(
                left=thick() if j == 1 else hair(),
                right=thick() if is_last_col else hair(),
                top=hair(), bottom=bot,
            )
            # colour t-stat by sign/significance
            if j == 4 and abs(tstat) > 1.96:
                c.font = Font(name="Calibri", size=10, bold=True,
                              color="375623" if tstat > 0 else RED_HEX)

        current_row += 1

    # ── empty spacer row ─────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 8
    for j in range(1, total_cols + 1):
        c = ws.cell(current_row, j, "")
        c.fill = PatternFill("solid", fgColor=WHITE)
        c.border = Border(
            left=thick() if j == 1 else hair(),
            right=thick() if j == total_cols else hair(),
        )
    current_row += 1

    # ── R² row ───────────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 15
    ws.merge_cells(f"A{current_row}:C{current_row}")
    c = ws.cell(current_row, 1, "R²")
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border = Border(left=thick(), right=hair(), top=thin(), bottom=thick())

    ws.merge_cells(f"D{current_row}:{get_column_letter(total_cols)}{current_row}")
    c = ws.cell(current_row, 4, f"{r2:.4f}")
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=hair(), right=thick(), top=thin(), bottom=thick())
    current_row += 1

    # ── model fit section ────────────────────────────────────────────────────
    ws.merge_cells(f"A{current_row}:{get_column_letter(total_cols)}{current_row}")
    c = ws.cell(current_row, 1, "MODEL FIT")
    c.font = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = Border(top=thick(), bottom=thin(), left=thick(), right=thick())
    ws.row_dimensions[current_row].height = 15
    current_row += 1

    fit_rows = [
        ("Observations",        str(n_obs)),
        ("R²",                  f"{r2:.4f}"),
        ("Adjusted R²",         f"{adj_r2:.4f}"),
        ("F-Statistic",         f"{f_stat:.2f}"),
        ("p(F-Statistic)",      f"{f_pval:.4f}"),
        ("Dep. Var. Mean (ann.)",f"{float(y.mean()) * 12:.2%}"),
        ("Dep. Var. Vol (ann.)", f"{float(y.std()) * np.sqrt(12):.2%}"),
    ]

    for i, (label, val) in enumerate(fit_rows):
        is_last = (i == len(fit_rows) - 1)
        bg = WHITE if i % 2 == 0 else ALT_ROW
        bot = thick() if is_last else hair()
        ws.row_dimensions[current_row].height = 15

        # label spans cols 1-2, value spans cols 3-6
        ws.merge_cells(f"A{current_row}:B{current_row}")
        c = ws.cell(current_row, 1, label)
        c.font = Font(name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border = Border(left=thick(), right=thin(), top=hair(), bottom=bot)

        ws.merge_cells(f"C{current_row}:{get_column_letter(total_cols)}{current_row}")
        c = ws.cell(current_row, 3, val)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=thin(), right=thick(), top=hair(), bottom=bot)

        current_row += 1

    # ── footnote ─────────────────────────────────────────────────────────────
    ws.row_dimensions[current_row].height = 5
    current_row += 1
    ws.merge_cells(f"A{current_row}:{get_column_letter(total_cols)}{current_row}")
    c = ws.cell(current_row, 1,
                "Daily factor returns compounded to monthly. "
                "Factors sourced from factor_returns.xlsx. "
                "Green rows = significant at 5% level. Bold t-stats = |t| > 1.96.")
    c.font = Font(name="Calibri", italic=True, size=8, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[current_row].height = 14

    # ── column widths ─────────────────────────────────────────────────────────
    col_widths = [22, 14, 12, 14, 12, 7]
    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "B5"

    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    print("Loading factor returns...")
    monthly_factors = load_monthly_factors()

    print("Loading strategy returns...")
    strategy = load_strategy_returns()

    print("Merging...")
    merged = strategy.join(monthly_factors, how="inner").dropna()
    print(f"  Overlapping sample: {merged.index.min().strftime('%b %Y')} "
          f"to {merged.index.max().strftime('%b %Y')} "
          f"({len(merged)} months)")

    y     = merged["ew_mkt_neutral"]
    X_raw = merged[FACTORS]

    print("Running OLS regression...")
    result = run_regression(y, X_raw)

    print("\n" + result.summary().as_text())

    out_path = OUT_DIR / "Factor_Regression.xlsx"
    build_excel(result, y, X_raw, out_path)


if __name__ == "__main__":
    main()
